#!/usr/bin/env python3
"""
Export 《策划案.xlsx》 into runtime JSON files for Data/Config module.

Default source:
  Assets\\ConfigSource\\策划案.xlsx
Default output:
  Assets\\Resources\\Configs\\PlanningXlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def as_int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return int(value)
    text = as_text(value)
    if text == "":
        return 0
    try:
        return int(float(text))
    except Exception:
        return 0


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = as_text(value).lower()
    return text in ("y", "yes", "true", "1")


def as_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = as_text(value)
    if text == "":
        return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0


def split_ids(value: Any) -> List[str]:
    text = as_text(value)
    if not text:
        return []
    parts = re.split(r"[;；,，\n\r]+", text)
    return [p.strip() for p in parts if p and p.strip()]


def sanitize_name(name: str) -> str:
    safe = re.sub(r'[<>:"/\\|?*]+', "_", name.strip())
    safe = re.sub(r"\s+", "_", safe)
    return safe or "sheet"


def unique_id(raw: str, seen: set[str], fallback_prefix: str, index: int) -> str:
    value = raw.strip()
    if not value:
        value = f"{fallback_prefix}_{index}"
    base = value
    seq = 2
    while value in seen:
        value = f"{base}_{seq}"
        seq += 1
    seen.add(value)
    return value


@dataclass
class Issue:
    level: str
    code: str
    message: str
    sheet: Optional[str] = None
    row: Optional[int] = None
    column: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "level": self.level,
            "code": self.code,
            "message": self.message,
            "sheet": self.sheet,
            "row": self.row,
            "column": self.column,
        }


@dataclass
class ExportResult:
    sheet_name: str
    output_file: str
    rows_exported: int
    issues: List[Issue] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "sheetName": self.sheet_name,
            "outputFile": self.output_file,
            "rowsExported": self.rows_exported,
            "issues": [x.to_dict() for x in self.issues],
        }


class RowAccessor:
    def __init__(self, headers: Dict[str, int], row_values: List[Any]):
        self.headers = headers
        self.row_values = row_values

    def get(self, name: str) -> Any:
        idx = self.headers.get(name)
        if idx is None:
            return None
        if idx < 0 or idx >= len(self.row_values):
            return None
        return self.row_values[idx]


def is_empty_row(values: List[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def load_openpyxl():
    try:
        import openpyxl  # type: ignore
    except Exception as ex:
        raise RuntimeError("Missing dependency openpyxl. Install: pip install openpyxl") from ex
    return openpyxl


Mapper = Callable[[RowAccessor, int, set[str], List[Issue], str], Optional[Dict[str, Any]]]


def map_core_framework(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    module_name = as_text(row.get("模块名称"))
    if not module_name:
        issues.append(Issue("warning", "row.skip.empty_module", "模块名称为空，已跳过。", sheet, row_index, "模块名称"))
        return None
    rid = unique_id(module_name, seen, "module", row_index)
    return {
        "id": rid,
        "moduleName": module_name,
        "stage": as_text(row.get("所在状态/Stage")),
        "featureDescription": as_text(row.get("主要功能描述")),
        "note": as_text(row.get("备注说明")),
    }


def map_character(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    rid = unique_id(as_text(row.get("角色ID")), seen, "character", row_index)
    return {
        "id": rid,
        "roleName": as_text(row.get("角色名称")),
        "initialHp": as_int(row.get("初始HP")),
        "initialMpSp": as_int(row.get("初始MP/SP")),
    }


def map_skill(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    rid = unique_id(as_text(row.get("技能ID")), seen, "skill", row_index)
    return {
        "id": rid,
        "ownerRoleId": as_text(row.get("所属角色/通用")),
        "quality": as_int(row.get("品质")),
        "singleUsePerBattle": as_bool(row.get("是否一场战斗只能使用一次（y/n）")),
        "skillName": as_text(row.get("技能名称")),
        "skillType": as_text(row.get("类型(主动/被动/BUFF)")),
        "costValue": as_int(row.get("消耗值")),
        "target": as_text(row.get("目标")),
        "effects": as_text(row.get("效果(伤害/附加状态)")),
        "description": as_text(row.get("详细描述")),
    }


def map_enemy(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    rid = unique_id(as_text(row.get("id")), seen, "enemy", row_index)
    elite_text = as_text(row.get("类型（是否精英？）")).lower()
    return {
        "id": rid,
        "isElite": elite_text in ("yes", "y", "精英", "true", "1"),
        "battlePattern": as_text(row.get("战斗出招")),
        "rewardId": as_text(row.get("战斗奖励")),
    }


def map_design_guideline(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    topic = as_text(row.get("项目"))
    content = as_text(row.get("整理后的口径"))
    if not topic and not content:
        return None
    rid = unique_id(topic or content, seen, "guideline", row_index)
    return {"id": rid, "topic": topic, "content": content}


def map_battle_unit(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    rid = unique_id(as_text(row.get("Id")), seen, "unit", row_index)
    return {
        "id": rid,
        "displayName": as_text(row.get("DisplayName")),
        "faction": as_text(row.get("Faction")),
        "role": as_text(row.get("Role")),
        "hp": as_float(row.get("HP")),
        "attack": as_float(row.get("Attack")),
        "defense": as_float(row.get("Defense")),
        "speed": as_float(row.get("Speed")),
        "mana": as_float(row.get("Mana")),
        "innateSkillIds": split_ids(row.get("InnateSkillIds")),
        "designNotes": as_text(row.get("DesignNotes")),
    }


def map_new_skill(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    rid = unique_id(as_text(row.get("Id")), seen, "skill", row_index)
    return {
        "id": rid,
        "displayName": as_text(row.get("DisplayName")),
        "ownerUnitId": as_text(row.get("OwnerUnitId")),
        "skillType": as_text(row.get("SkillType")),
        "targetType": as_text(row.get("TargetType")),
        "level": as_int(row.get("Level")),
        "manaCost": as_int(row.get("ManaCost")),
        "oncePerBattle": as_bool(row.get("OncePerBattle")),
        "quality": as_int(row.get("Quality")),
        "canCastConditions": as_text(row.get("CanCastConditions")),
        "applyActions": as_text(row.get("ApplyActions")),
        "animationCue": as_text(row.get("AnimationCue")),
        "sfxCue": as_text(row.get("SfxCue")),
        "designNotes": as_text(row.get("DesignNotes")),
    }


def map_battle_effect(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    rid = unique_id(as_text(row.get("Id")), seen, "effect", row_index)
    return {
        "id": rid,
        "displayName": as_text(row.get("DisplayName")),
        "effectType": as_text(row.get("EffectType")),
        "statusType": as_text(row.get("StatusType")),
        "initialTurns": as_int(row.get("InitialTurns")),
        "maxStackCount": as_int(row.get("MaxStackCount")),
        "stackRule": as_text(row.get("StackRule")),
        "triggerTiming": as_text(row.get("TriggerTiming")),
        "applyActions": as_text(row.get("ApplyActions")),
        "visualCue": as_text(row.get("VisualCue")),
        "sfxCue": as_text(row.get("SfxCue")),
        "designNotes": as_text(row.get("DesignNotes")),
    }


def map_battle_reward_new(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    rid = unique_id(as_text(row.get("Id")), seen, "reward", row_index)
    return {
        "id": rid,
        "displayName": as_text(row.get("DisplayName")),
        "rewardRarity": as_text(row.get("RewardRarity")),
        "applyActions": as_text(row.get("ApplyActions")),
        "designNotes": as_text(row.get("DesignNotes")),
    }


def map_battle_reward(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    reward_name = as_text(row.get("奖励类型"))
    rid = unique_id(reward_name, seen, "reward", row_index)
    options = [
        as_text(row.get("奖励选项1")),
        as_text(row.get("奖励选项2")),
        as_text(row.get("奖励选项3")),
        as_text(row.get("奖励选项4")),
        as_text(row.get("奖励选项5")),
    ]
    options = [x for x in options if x]
    return {"id": rid, "rewardName": reward_name, "options": options}


def map_state(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    raw_id = as_text(row.get("State ID"))
    if not raw_id:
        raw_id = as_text(row.get("状态名称"))
    rid = unique_id(raw_id, seen, "state", row_index)
    return {
        "id": rid,
        "stateName": as_text(row.get("状态名称")),
        "stateType": as_text(row.get("类型(增益/减益/控制)")),
        "duration": as_text(row.get("持续时长")),
        "affectedAttribute": as_text(row.get("作用属性")),
        "valueDescription": as_text(row.get("数值/百分比")),
        "description": as_text(row.get("状态描述")),
    }


def map_design_note(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    content = as_text(row.get("规则说明"))
    if not content:
        return None
    rid = unique_id("", seen, "rule_note", row_index)
    return {"id": rid, "content": content}


def map_formula(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    raw = as_text(row.get("公式类型"))
    if not raw and is_empty_row([row.get("公式类型"), row.get("计算公式内容"), row.get("涉及变量"), row.get("附加特效触发概率"), row.get("说明")]):
        return None
    rid = unique_id(raw, seen, "formula", row_index)
    return {
        "id": rid,
        "formulaType": raw,
        "formulaContent": as_text(row.get("计算公式内容")),
        "variables": as_text(row.get("涉及变量")),
        "effectTriggerChance": as_text(row.get("附加特效触发概率")),
        "note": as_text(row.get("说明")),
    }


def map_item(row: RowAccessor, row_index: int, seen: set[str], issues: List[Issue], sheet: str):
    rid = unique_id(as_text(row.get("物品ID")), seen, "item", row_index)
    return {
        "id": rid,
        "name": as_text(row.get("名称")),
        "itemType": as_text(row.get("类型")),
        "propertyBonus1": as_text(row.get("属性加成1")),
        "propertyBonus2": as_text(row.get("属性加成2")),
        "source": as_text(row.get("获取途径")),
        "restriction": as_text(row.get("限制条件")),
    }


SPECS = [
    {
        "sheetNames": ["设计口径"],
        "output": "design_guidelines.json",
        "headerAliases": {},
        "mapper": map_design_guideline,
    },
    {
        "sheetNames": ["BattleUnit"],
        "output": "battle_units.json",
        "headerAliases": {},
        "mapper": map_battle_unit,
    },
    {
        "sheetNames": ["Skill"],
        "output": "skills.json",
        "headerAliases": {},
        "mapper": map_new_skill,
    },
    {
        "sheetNames": ["BattleEffect"],
        "output": "battle_effects.json",
        "headerAliases": {},
        "mapper": map_battle_effect,
    },
    {
        "sheetNames": ["BattleReward"],
        "output": "battle_rewards.json",
        "headerAliases": {},
        "mapper": map_battle_reward_new,
    },
]


def find_spec(sheet_name: str):
    for spec in SPECS:
        if sheet_name in spec["sheetNames"]:
            return spec
    return None


def read_headers(row_values: List[Any], aliases: Dict[str, str]) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for idx, cell in enumerate(row_values):
        key = as_text(cell)
        if not key:
            continue
        mapped = aliases.get(key, key)
        if mapped and mapped not in headers:
            headers[mapped] = idx
    return headers


def export_sheet(ws, output_dir: Path) -> ExportResult:
    sheet = ws.title
    issues: List[Issue] = []
    spec = find_spec(sheet)
    if spec is None:
        file_name = f"{sanitize_name(sheet)}.json"
        issues.append(Issue("warning", "sheet.unsupported", "No export spec for this sheet; exported as empty.", sheet))
        payload = {"items": []}
        (output_dir / file_name).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return ExportResult(sheet, file_name, 0, issues)

    file_name = spec["output"]

    if sheet == "战斗奖励":
        seen_ids: set[str] = set()
        items: List[Dict[str, Any]] = []
        for row_index in range(1, ws.max_row + 1):
            row_values = [ws.cell(row_index, c).value for c in range(1, ws.max_column + 1)]
            if is_empty_row(row_values):
                continue
            reward_name = as_text(row_values[0])
            if not reward_name:
                continue
            rid = unique_id(reward_name, seen_ids, "reward", row_index)
            options = [as_text(v) for v in row_values[1:] if as_text(v)]
            items.append({"id": rid, "rewardName": reward_name, "options": options})
        if len(items) == 0:
            issues.append(Issue("warning", "sheet.no_data", "No data rows exported.", sheet))
        (output_dir / file_name).write_text(json.dumps({"items": items}, ensure_ascii=False, indent=2), encoding="utf-8")
        return ExportResult(sheet, file_name, len(items), issues)

    if sheet == "一些混淆的点":
        seen_ids: set[str] = set()
        items: List[Dict[str, Any]] = []
        for row_index in range(1, ws.max_row + 1):
            text = as_text(ws.cell(row_index, 1).value)
            if not text:
                continue
            rid = unique_id("", seen_ids, "rule_note", row_index)
            items.append({"id": rid, "content": text})
        if len(items) == 0:
            issues.append(Issue("warning", "sheet.no_data", "No data rows exported.", sheet))
        (output_dir / file_name).write_text(json.dumps({"items": items}, ensure_ascii=False, indent=2), encoding="utf-8")
        return ExportResult(sheet, file_name, len(items), issues)

    aliases = spec["headerAliases"]
    mapper: Mapper = spec["mapper"]

    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = read_headers(header_row, aliases)

    if not headers:
        issues.append(Issue("error", "header.empty", "No usable headers in first row.", sheet, 1))
        (output_dir / file_name).write_text(json.dumps({"items": []}, ensure_ascii=False, indent=2), encoding="utf-8")
        return ExportResult(sheet, file_name, 0, issues)

    seen_ids: set[str] = set()
    items: List[Dict[str, Any]] = []
    for row_index in range(2, ws.max_row + 1):
        row_values = [ws.cell(row_index, c).value for c in range(1, ws.max_column + 1)]
        if is_empty_row(row_values):
            continue
        row = RowAccessor(headers, row_values)
        mapped = mapper(row, row_index, seen_ids, issues, sheet)
        if mapped is None:
            continue
        items.append(mapped)

    if len(items) == 0:
        issues.append(Issue("warning", "sheet.no_data", "No data rows exported.", sheet))

    out_path = output_dir / file_name
    out_path.write_text(json.dumps({"items": items}, ensure_ascii=False, indent=2), encoding="utf-8")
    return ExportResult(sheet, file_name, len(items), issues)


def write_text(path: Path, lines: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Export planning xlsx to runtime json.")
    parser.add_argument("--source", default=r"Assets\ConfigSource\策划案.xlsx")
    parser.add_argument("--output", default=r"Assets\Resources\Configs\PlanningXlsx")
    parser.add_argument("--log", default=r"TestLogs\ConfigExport\planning_xlsx_export.log")
    parser.add_argument("--manifest", default=r"TestLogs\ConfigExport\planning_xlsx_export_manifest.json")
    parser.add_argument("--fail-on-warning", action="store_true")
    args = parser.parse_args()

    source = Path(args.source)
    output = Path(args.output)
    log_path = Path(args.log)
    manifest_path = Path(args.manifest)

    now = datetime.now().isoformat(timespec="seconds")
    logs = [f"[INFO] Export start: {now}", f"[INFO] Source: {source}", f"[INFO] Output: {output}"]

    if not source.exists():
        logs.append(f"[ERROR] source.missing: {source}")
        write_text(log_path, logs)
        manifest_path.parent.mkdir(parents=True, exist_ok=True)
        manifest_path.write_text(
            json.dumps(
                {"success": False, "source": str(source), "output": str(output), "issues": [{"code": "source.missing"}]},
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        return 1

    openpyxl = load_openpyxl()
    wb = openpyxl.load_workbook(source, data_only=True)
    output.mkdir(parents=True, exist_ok=True)

    results: List[ExportResult] = []
    all_issues: List[Issue] = []
    for name in wb.sheetnames:
        result = export_sheet(wb[name], output)
        results.append(result)
        all_issues.extend(result.issues)
        logs.append(f"[INFO] {result.sheet_name} -> {result.output_file}, rows={result.rows_exported}")

    err = sum(1 for i in all_issues if i.level == "error")
    warn = sum(1 for i in all_issues if i.level == "warning")
    for issue in all_issues:
        logs.append(
            f"[{issue.level.upper()}] {issue.code} sheet={issue.sheet} row={issue.row} col={issue.column} msg={issue.message}"
        )
    success = err == 0 and (warn == 0 if args.fail_on_warning else True)
    logs.append(f"[INFO] Export done. errors={err}, warnings={warn}, success={success}")
    write_text(log_path, logs)

    manifest = {
        "source": str(source),
        "output": str(output),
        "exportedAt": now,
        "success": success,
        "errorCount": err,
        "warningCount": warn,
        "issues": [i.to_dict() for i in all_issues],
        "sheets": [r.to_dict() for r in results],
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return 0 if success else 2


if __name__ == "__main__":
    sys.exit(main())

